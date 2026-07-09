from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from urllib.parse import urlparse
import re


def generate_threat_report(sessions):

    wb = Workbook()

    # Create worksheets
    ws_ip = wb.active
    ws_ip.title = "IP"

    ws_domain = wb.create_sheet("Domain")
    ws_hash = wb.create_sheet("Hashes")
    ws_url = wb.create_sheet("URL")

    # Header Style
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    font = Font(bold=True)

    worksheets = [
        (ws_ip, "IP"),
        (ws_domain, "Domain"),
        (ws_hash, "Hashes"),
        (ws_url, "URL"),
    ]

    for ws, title in worksheets:
        ws["A1"] = title
        ws["A1"].fill = fill
        ws["A1"].font = font
        ws.column_dimensions["A"].width = 80

    ips = set()
    domains = set()
    urls = set()
    hashes = set()

    for session in sessions.values():

        ip = session.get("src_ip")

        if ip:
            ips.add(ip)

        commands = session.get("commands", [])

        for cmd in commands:

            # Extract URLs
            found_urls = re.findall(r'https?://\S+', cmd)

            for url in found_urls:

                urls.add(url)

                try:
                    domains.add(urlparse(url).netloc)
                except Exception:
                    pass

            # Extract SHA256 Hashes
            found_hashes = re.findall(r'\b[a-fA-F0-9]{64}\b', cmd)

            for h in found_hashes:
                hashes.add(h)

    # Write IPs
    row = 2
    for ip in sorted(ips):
        ws_ip[f"A{row}"] = ip
        row += 1

    # Write Domains
    row = 2
    for domain in sorted(domains):
        ws_domain[f"A{row}"] = domain
        row += 1

    # Write Hashes
    row = 2
    for h in sorted(hashes):
        ws_hash[f"A{row}"] = h
        row += 1

    # Write URLs
    row = 2
    for url in sorted(urls):
        ws_url[f"A{row}"] = url
        row += 1

    wb.save("output/threat_intelligence.xlsx")

    print("Threat Intelligence Excel saved to output/threat_intelligence.xlsx")